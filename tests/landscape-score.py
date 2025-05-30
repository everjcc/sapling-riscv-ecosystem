import pandas as pd
import numpy as np
from datetime import datetime
import argparse
from openpyxl import load_workbook

param_name = 'performance-parameters'
benchmark = 'Benchmark'


required_params_info = {
    'case1': ['duration_time(ns)', 'instructions', 'L1-dcache-load-misses'],
    'case2': ['duration time', 'instructions', 'L1 dcache load misses'],
    'case3': ['duration_time', 'instructions', 'L1-dcache-load-misses']
}

to_number = lambda x: 0 if pd.isna(x) else x


def find_and_count_functionality(input_file, sheet_name):
    df = pd.read_excel(input_file, sheet_name=sheet_name, header=1)

    benchmark_total_count = len(df)

    column_2_count = df.iloc[:, 1].eq('Y').sum()
    column_4_count = df.iloc[:, 3].eq('Y').sum()
    both_columns_count = ((df.iloc[:, 1] == 'Y') & (df.iloc[:, 3] == 'Y')).sum()

    a1 = column_2_count / len(df)
    a2 = column_4_count / len(df)


    if a1 == 0:
        raise ValueError("a1 cannot be 0") 

    if a2 == 0:
        return 0, column_2_count, column_4_count, both_columns_count, a1, a2

    ratio = a2 / a1
    score_w2 = 20 * ratio if ratio < 1 else 20

    return benchmark_total_count, score_w2, column_2_count, column_4_count, both_columns_count, a1, a2



def calculate_prosperity_score(input_file, sheet_name):
    df = pd.read_excel(input_file, sheet_name=sheet_name)
    if df.size == 0:
        return 0
    riscv_sum = to_number(df.iloc[5, 0])
    arm_sum = to_number(df.iloc[5, 1])
    riscv_q1 = to_number(df.iloc[0, 1])
    arm_q1 = to_number(df.iloc[1, 1])
    riscv_q2 = to_number(df.iloc[0, 2])
    arm_q2 = to_number(df.iloc[1, 2])
    riscv_q3 = to_number(df.iloc[0, 3])
    arm_q3 = to_number(df.iloc[1, 3])
    riscv_q4 = to_number(df.iloc[0, 4])
    arm_q4 = to_number(df.iloc[1, 4])

    b22 = riscv_q1 + riscv_q2 + riscv_q3 + riscv_q4
    b12 = arm_q1 + arm_q2 + arm_q3 + arm_q4
    b21 = riscv_sum + b22
    b11 = arm_sum + b12
 
    
    if b11 == 0:
        landscape_score_w3_t = 6
        paper_score_w3_t = 3
    
    else:
        b21_b11_ratio = b21 / b11
        landscape_score_w3_t = 6 * b21_b11_ratio if b21_b11_ratio < 1 else 6
        paper_score_w3_t = 3 * b21_b11_ratio if b21_b11_ratio < 1 else 3
        
    if b12 == 0:
        landscape_score_w3_2024 = 14
        paper_score_w3_2024 = 7
    
    else:
        b22_b12_ratio = b22 / b12
        landscape_score_w3_2024 = 14 * b22_b12_ratio if b22_b12_ratio < 1 else 14
        paper_score_w3_2024 = 7 * b22_b12_ratio if b22_b12_ratio < 1 else 7
    
    landscape_score_w3 =landscape_score_w3_t + landscape_score_w3_2024
    paper_score_w3 = paper_score_w3_t + paper_score_w3_2024
        
    return landscape_score_w3, paper_score_w3, b11, b21, b12, b22

def calculate_heterogeneous_support_score(input_file, sheet_name):
    df = pd.read_excel(input_file, sheet_name=sheet_name)

    column_2_total_count = len(df.iloc[:, 1])

    column_2_y_count = df.iloc[:, 1].eq('Y').sum()

    column_3_total_count = len(df.iloc[:, 2])

    column_3_y_count = df.iloc[:, 2].eq('Y').sum()
    c1 = column_2_y_count / column_2_total_count if column_2_total_count != 0 else 0
    c2 = column_3_y_count / column_3_total_count if column_3_total_count != 0 else 0

    if c1 == 0:
        landscape_score_w4 = 20
        paper_score_w4 = 10
    else:
        ratio_c2_c1 = c2 / c1
        landscape_score_w4 = 20 * ratio_c2_c1 if ratio_c2_c1 < 1 else 20
        paper_score_w4 = 10 * ratio_c2_c1 if ratio_c2_c1 < 1 else 10

    return landscape_score_w4, paper_score_w4

def calculate_performance_score(input_file, sheet_name):
    df = pd.read_excel(input_file, sheet_name=sheet_name, header=1)

    for name, required_params in required_params_info.items():
        match_info = df[param_name].isin(required_params)


        flag = False
        for param in required_params:
            flag |= (not any(df[param_name].isin([param])))
        if flag:
            continue

        #print(f'{name}:')

        df_filtered = df[match_info]

        pivot_df = df_filtered.pivot_table(index=benchmark, columns=param_name, values=['ARM（kunpeng920-4826)', 'RISC-V（SOPHON SG2042）'])
        pivot_df.columns = ['_'.join(col).strip() for col in pivot_df.columns.values]

        pivot_df['duration_ratio'] = pivot_df[f'RISC-V（SOPHON SG2042）_{required_params[0]}'] / pivot_df[f'ARM（kunpeng920-4826)_{required_params[0]}']
        pivot_df['instructions_ratio'] = pivot_df[f'RISC-V（SOPHON SG2042）_{required_params[1]}'] / pivot_df[f'ARM（kunpeng920-4826)_{required_params[1]}']
        pivot_df['load_misses_ratio'] = pivot_df[f'RISC-V（SOPHON SG2042）_{required_params[2]}'] / pivot_df[f'ARM（kunpeng920-4826)_{required_params[2]}']

        break


    pivot_df = pivot_df.dropna(subset=['duration_ratio', 'instructions_ratio', 'load_misses_ratio'])

    q = len(pivot_df)

    ST = np.prod(pivot_df['duration_ratio'] ** (1 / q))
    SN = np.prod(pivot_df['instructions_ratio'] ** (1 / q))
    SM = np.prod(pivot_df['load_misses_ratio'] ** (1 / q))


    inverse_pa = 1 / ST
    inverse_cq = 1/ SN
    inverse_mp = 1 / SM
    score_pa = 10 * inverse_pa if inverse_pa < 1 else 10
    score_cq = 5 * inverse_cq if inverse_cq < 1 else 5
    score_mp = 5 * inverse_mp if inverse_mp < 1 else 5
    score_tatol = score_pa + score_cq + score_mp

    paper_score_pa = 20 * inverse_pa if inverse_pa < 1 else 20
    paper_score_cq = 20 * inverse_cq if inverse_cq < 1 else 20
    paper_score_mp = 20 * inverse_mp if inverse_mp < 1 else 20
    '''print(q)
    print(ST)
    print(SN)
    print(SM)
    print(s)
    print(inverse_s)'''
    return paper_score_pa, paper_score_cq, paper_score_mp, score_tatol

def append_to_excel(data_frame, output_file, sheet_name):
    try:
   
        writer = pd.ExcelWriter(output_file, engine='openpyxl', mode='a')
        writer.book = load_workbook(output_file)

        if sheet_name in writer.book.sheetnames:
            existing_data = pd.read_excel(output_file, sheet_name=sheet_name)
           
            combined_data = pd.concat([existing_data, data_frame], ignore_index=True)
        else:
          
            combined_data = data_frame

        combined_data.to_excel(writer, sheet_name=sheet_name, index=False)

        writer.save()
        writer.close()
    except FileNotFoundError:

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            data_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    # For pandas version > 1.2.x
    except AttributeError:
 
        existing_data = pd.read_excel(output_file, sheet_name=sheet_name)
      
        combined_data = pd.concat([existing_data, data_frame], ignore_index=True)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            combined_data.to_excel(writer, sheet_name=sheet_name, index=False)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--name', help='input software name', default='Kafka')
    package_name = parser.parse_args().name

    current_date = datetime.now().date()

    formatted_date = current_date.strftime('%m%d')
    input_file = f'{package_name}-Landscape.xlsx'
    output_file_xlsx = "./output_result/" + f'{package_name}-output-{formatted_date}.xlsx'

    output_file_txt = "./output_result/" + f'{package_name}-output-{formatted_date}.txt'

    d_r, c_r, l_m_r, performance_score = calculate_performance_score(input_file, 'performance')
    benchmark_total_count, functionality_score, column_2_count, column_4_count, both_columns_count, a1, a2 = find_and_count_functionality(input_file, 'functionality')

    landscape_prosperity_score, paper_prosperity_score, arm_total, riscv_total, arm_2024, riscv_2024  = calculate_prosperity_score(input_file, 'ODA')
    landscape_heterogeneous_support_score, paper_heterogeneous_support_score = calculate_heterogeneous_support_score(input_file, 'HCS')

  

    data_info = {
        f'包名-日期': f'{package_name}-{formatted_date}',
        f'PA': f'{d_r:.2f}',
        f'CQ': f'{c_r:.2f}',
        f'MP': f'{l_m_r:.2f}',
        f'paper-functionality (w4)': f'{functionality_score:.2f}',
        f'paper-ODA(w5)': f'{paper_prosperity_score}',
        f'paper-HCS(w6)': f'{paper_heterogeneous_support_score}',
    }

    append_to_excel(pd.DataFrame({k: [v] for k, v in data_info.items()}),
                    f'output.xlsx',
                    'output')
